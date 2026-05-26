"""
EcoPulse Excel Audit Generator
Creates a professional multi-sheet audit workbook from the cleaned dataset.
"""

import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from sqlalchemy import create_engine

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DATA_DIR  = os.path.join(BASE_DIR, '..', 'data')
DB_PATH   = os.path.abspath(os.path.join(DATA_DIR, 'ecopulse.db'))
OUT_PATH  = os.path.abspath(os.path.join(DATA_DIR, 'EcoPulse_Audit.xlsx'))


HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
ACCENT_FILL  = PatternFill("solid", fgColor="D6E4F0")
TITLE_FONT   = Font(bold=True, size=14, color="1F4E79")
THIN_BORDER  = Border(
    left=Side(style='thin', color='AAAAAA'),
    right=Side(style='thin', color='AAAAAA'),
    top=Side(style='thin', color='AAAAAA'),
    bottom=Side(style='thin', color='AAAAAA')
)


def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(length + 2, min_w), max_w)


def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def write_cover(wb, df):
    ws = wb.create_sheet("Cover")
    wb.active = ws

    ws['B2'] = "EcoPulse: Smart City Energy Consumption"
    ws['B2'].font = Font(bold=True, size=18, color="1F4E79")
    ws['B3'] = "Data Audit Workbook"
    ws['B3'].font = Font(size=13, color="444444")
    ws['B5'] = f"Dataset: PJME Hourly Energy Consumption"
    ws['B6'] = f"Records: {len(df):,}"
    ws['B7'] = f"Date Range: {df['datetime'].min().date()} to {df['datetime'].max().date()}"
    ws['B8'] = f"Columns: {', '.join(df.columns[:6].tolist())} ..."
    ws['B10'] = "Data Quality Summary"
    ws['B10'].font = Font(bold=True, size=12, color="1F4E79")

    issues = [
        ("Missing Values (PJME_MW)", df['PJME_MW'].isna().sum()),
        ("Duplicate Rows",           df.duplicated(subset=['datetime']).sum()),
        ("Zero/Negative MW Rows",    (df['PJME_MW'] <= 0).sum()),
        ("Outliers (>3 std)",        ((df['PJME_MW'] - df['PJME_MW'].mean()).abs() >
                                     3 * df['PJME_MW'].std()).sum()),
    ]
    for i, (label, val) in enumerate(issues, start=11):
        ws[f'B{i}'] = label
        ws[f'C{i}'] = int(val)
        ws[f'D{i}'] = "✅ Clean" if val == 0 else "⚠️ Review"

    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.sheet_view.showGridLines = False


def write_raw_sample(wb, df):
    ws = wb.create_sheet("Raw Data Sample")
    sample = df[['datetime', 'PJME_MW', 'year', 'month', 'hour',
                  'is_weekend']].head(500).copy()
    headers = list(sample.columns)
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for _, row in sample.iterrows():
        ws.append(row.tolist())
    auto_width(ws)


def write_summary_stats(wb, df):
    ws = wb.create_sheet("Summary Statistics")
    ws['A1'] = "Descriptive Statistics — PJME_MW"
    ws['A1'].font = TITLE_FONT

    stats = df['PJME_MW'].describe(percentiles=[0.1, 0.25, 0.5, 0.75, 0.9])
    extra = pd.Series({
        'skewness': df['PJME_MW'].skew(),
        'kurtosis': df['PJME_MW'].kurt(),
        'std_dev':  df['PJME_MW'].std(),
    })
    stats = pd.concat([stats, extra])

    ws.append(["Metric", "Value"])
    style_header_row(ws, 2, 2)
    for stat_name, val in stats.items():
        ws.append([stat_name, round(float(val), 2)])

    ws['A15'] = "Monthly Averages"
    ws['A15'].font = TITLE_FONT
    monthly = df.groupby('month')['PJME_MW'].agg(['mean', 'max', 'min']).round(1)
    monthly.columns = ['Avg MW', 'Peak MW', 'Min MW']
    monthly = monthly.reset_index()
    headers = list(monthly.columns)
    ws.append(headers)
    style_header_row(ws, 16, len(headers))
    for _, row in monthly.iterrows():
        ws.append(row.tolist())

    auto_width(ws)


def write_pivot_analysis(wb, df):
    ws = wb.create_sheet("Pivot Analysis")
    ws['A1'] = "Average MW by Hour of Day"
    ws['A1'].font = TITLE_FONT

    hourly = df.groupby('hour')['PJME_MW'].mean().round(1).reset_index()
    hourly.columns = ['Hour', 'Avg MW']
    ws.append(list(hourly.columns))
    style_header_row(ws, 2, 2)
    for _, row in hourly.iterrows():
        ws.append(row.tolist())

    # Line chart
    chart = LineChart()
    chart.title = "Average Energy Demand by Hour"
    chart.style = 10
    chart.y_axis.title = "MW"
    chart.x_axis.title = "Hour"
    data_ref   = Reference(ws, min_col=2, min_row=2, max_row=26)
    labels_ref = Reference(ws, min_col=1, min_row=3, max_row=26)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    chart.width  = 18
    chart.height = 12
    ws.add_chart(chart, "D2")

    ws['A30'] = "Average MW by Month"
    ws['A30'].font = TITLE_FONT
    monthly = df.groupby('month')['PJME_MW'].mean().round(1).reset_index()
    monthly.columns = ['Month', 'Avg MW']
    ws.append(list(monthly.columns))
    style_header_row(ws, 31, 2)
    for _, row in monthly.iterrows():
        ws.append(row.tolist())

    auto_width(ws)


def write_yoy_analysis(wb, df):
    ws = wb.create_sheet("Year-over-Year")
    ws['A1'] = "Year-over-Year Analysis"
    ws['A1'].font = TITLE_FONT

    yoy = df.groupby('year')['PJME_MW'].agg(['mean', 'max', 'sum']).round(1)
    yoy.columns = ['Avg MW', 'Peak MW', 'Total MWh']
    yoy['YoY Change %'] = yoy['Avg MW'].pct_change().mul(100).round(2)
    yoy = yoy.reset_index()

    ws.append(list(yoy.columns))
    style_header_row(ws, 2, len(yoy.columns))
    for _, row in yoy.iterrows():
        ws.append(row.tolist())

    auto_width(ws)


def generate_excel_audit():
    print("Loading data from SQLite...")
    engine = create_engine(f'sqlite:///{DB_PATH}')
    df = pd.read_sql('SELECT * FROM energy_hourly ORDER BY datetime', engine)
    df['datetime'] = pd.to_datetime(df['datetime'])

    print("Building Excel workbook...")
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    write_cover(wb, df)
    write_raw_sample(wb, df)
    write_summary_stats(wb, df)
    write_pivot_analysis(wb, df)
    write_yoy_analysis(wb, df)

    wb.save(OUT_PATH)
    print(f"Excel audit saved: {OUT_PATH}")
    return OUT_PATH


if __name__ == '__main__':
    generate_excel_audit()
